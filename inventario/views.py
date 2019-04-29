from openpyxl import Workbook 
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from django.shortcuts import render
from django.core.paginator import Paginator
from django.views import View
from django.views.generic import DetailView
from django.views.generic.base import ContextMixin, TemplateView
from django.views.generic.edit import CreateView
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.http import JsonResponse, Http404, HttpResponse, HttpResponseForbidden
from django.utils.timezone import localtime, now


class ReportView(TemplateView):
    def append_filter(self, filter_, ws):
        for key, field in filter_.form.fields.items():
            value = filter_.data[key]
            if value is '':
                display_value = ''
            else:
                try:
                    display_value = dict(field.widget.choices)[value]
                except:
                    display_value = value
            ws.append(['{}: {}'.format(field.label, display_value)])
        ws.append([])


    def get(self, request, *args, **kwargs):
        if 'excel' in request.GET:
            wb = Workbook()
            ws = wb.active

            datetime = localtime(now())

            ws.append([self.title])
            ws.append(['Fecha de generación de reporte: {}'.format(datetime.strftime('%m-%d-%Y %I:%M %p'))])
            ws.append([])

            self.append_objects(request, ws, **kwargs)

            for column in ws.columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    if cell.coordinate in ws.merged_cells: 
                        continue
                    try: 
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_name].width = adjusted_width

            response = HttpResponse(content_type = 'application/ms-excel')
            content = 'attachment; filename = {0}'.format('{}_{}_{}'.format(self.filename, datetime.strftime('%m-%d-%Y_%I:%M-%p'), '.xlsx'))
            response['Content-Disposition'] = content
            wb.save(response)

            return response
        return Http404


class CreateToDetailView(PermissionRequiredMixin, CreateView):
    permission_required = ()
    change_permission = None

    def get_success_url(self):
        if self.change_permission is None or self.request.user.has_perm(self.change_permission):
            return self.object.get_absolute_url()
        return self.request.path_info


class DetailFormsView(PermissionRequiredMixin, DetailView):
    template_name = 'productos/base_detalle.html'
    permission_required = ()
    add_permission = None

    def get_context_data(self, **kwargs):
        self.object = self.get_object()
        context = super().get_context_data(**kwargs)

        if self.add_permission is None or self.request.user.has_perm(self.add_permission):
            context['can_add'] = True

        self.get_forms(context)

        return context


class ListView(PermissionRequiredMixin, ContextMixin, View):
    permission_required = ()
    add_permission = None
    change_permission = None
    post_form_list = []
    delete_form_list = []

    def get_context_data(self, **kwargs):
        context = super(ListView, self).get_context_data()

        if self.add_permission is not None and not self.request.user.has_perm(self.add_permission):
            context['can_add'] = False

        context['can_view'] = True

        context['filter_form'] = self.filter_(request=self.request, **kwargs).form

        return context

    def get_object_forms(self, obj):
        return {}

    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            page = request.GET.get('page')

            filter_ = self.filter_(request.GET, request=request, **kwargs)
            object_list = filter_.qs

            if page:
                paginator = Paginator(object_list, self.paginate_by)
                object_list = paginator.get_page(page).object_list
                page_count = paginator.num_pages
            else:
                page_count = 1
                
            forms = {}
            for obj in object_list:
                forms[obj.id] = self.get_object_forms(obj)
                
            context = {'object_list': object_list, 'object_forms_list': forms}

            if self.change_permission is None or self.request.user.has_perm(self.change_permission):
                context['can_change'] = True

            tabla_body = render(request, self.template_table, context)
            filter_form = render(request, 'busqueda_form.html', {'form': filter_.form})
            
            return JsonResponse({'results': tabla_body.content.decode("utf-8") , 'page_count': page_count, 'filter_form': filter_form.content.decode("utf-8")})

        return render(request, self.template_name, self.get_context_data(**kwargs))

    def post_forms(self, request):
        for post_form in self.post_form_list:
            if post_form.nombre_form in request.POST:
                form = post_form(request.POST)
                if form.is_valid():
                    form.save()
                    return HttpResponse('Formulario procesado exitosamente')
                return HttpResponse('Formulario no procesado')
        for delete_form in self.delete_form_list:
            if delete_form.nombre_form in request.POST:
                form = delete_form(request.POST)
                if form.is_valid():
                    form.delete()
                    return HttpResponse('Formulario procesado exitosamente')
                return HttpResponse('Formulario no procesado')

        return None

    def post(self, request):
        if request.is_ajax():
            response = self.post_forms(request)

            if response:
                return response

        return render(request, self.template_name, self.get_context_data())


class CreateUpdateView(PermissionRequiredMixin, ContextMixin, View):
    add_permission = None
    change_permission = None

    def has_permission(self):
        user = self.request.user
        return self.add_permission is None or self.change_permission is None or user.has_perm(self.add_permission) or user.has_perm(self.change_permission)

    def get_object(self, request):
        return None

    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            context = super().get_context_data()
            context['forms'] = []

            instance = self.get_object(request)

            user = request.user

            if instance is None:
                if self.add_permission is None or user.has_perm(self.add_permission):
                    form = self.form_class(request.POST or None)

                    if form.is_valid():
                        form.save()

                        context['forms'].append(self.form_class())
                        context['success'] = self.created_message
                    else:
                        context['forms'].append(form)
                else:
                    context['forms'].append(self.form_class())
                    context['success'] = 'No tiene permiso para crear este elemento.'
            else:
                if self.change_permission is None or user.has_perm(self.change_permission):
                    form = self.form_class(request.POST or None, instance=instance)

                    if form.is_valid():
                        form.save()

                        context['forms'].append(self.form_class())
                        context['success'] = self.updated_message
                    else:
                        context['forms'].append(form)
                else:
                    context['forms'].append(self.form_class())
                    context['success'] = 'No tiene permiso para actualizar este elemento.'

            return render(request, self.template_name, context)
        raise Http404


class CreateUpdateListView(ListView, ContextMixin):
    add_permission = None
    change_permission = None
    view_permission = None
    template_name = 'base_lista_crear_actualizar.html'
    template_form = 'forms.html'
    form_class = None

    def has_permission(self):
        user = self.request.user
        return self.add_permission is None or self.change_permission is None or self.view_permission is None or user.has_perm(self.add_permission) or user.has_perm(self.change_permission) or user.has_perm(self.view_permission)

    def get_object(self, request):
        return None

    def get_context_data(self):
        context = super().get_context_data()
        user = self.request.user

        if self.form_class is not None:
            if self.add_permission is None or self.change_permission is None or user.has_perm(self.add_permission) or user.has_perm(self.change_permission):
                context['forms'] = [self.form_class()]
        if self.view_permission is None or user.has_perm(self.view_permission):
            context['can_view'] = True
        else:
            context['can_view'] = False

        return context

    def get(self, request, *args, **kwargs):
        user = self.request.user

        if request.is_ajax():
            if self.view_permission is None or user.has_perm(self.view_permission): 
                return super().get(request, *args, **kwargs)
            return HttpResponseForbidden()
        
        return render(request, self.template_name, self.get_context_data())

    def post(self, request, *args, **kwargs):
        context = self.get_context_data()
        context['forms'] = []

        user = self.request.user

        if self.view_permission is None or user.has_perm(self.view_permission):
            context['can_view'] = True

            response = self.post_forms(request)

            if response:
                return response

        instance = self.get_object(request)

        if instance is None:
            if self.add_permission is None or user.has_perm(self.add_permission):
                form = self.form_class(request.POST or None)

                if form.is_valid():
                    form.save()

                    context['forms'].append(self.form_class())
                    context['success'] = self.created_message
                else:
                    context['forms'].append(form)
            else:
                context['forms'].append(self.form_class())
                context['success'] = 'No tiene permiso para crear este elemento.'
        else:
            if self.change_permission is None or user.has_perm(self.change_permission):
                form = self.form_class(request.POST or None, instance=instance)

                if form.is_valid():
                    form.save()

                    context['forms'].append(self.form_class())
                    context['success'] = self.updated_message
                else:
                    context['forms'].append(form)
            else:
                context['forms'].append(self.form_class())
                context['success'] = 'No tiene permiso para actualizar este elemento.'

        if request.is_ajax():
            return render(request, self.template_form, context)
        return render(request, self.template_name, context)