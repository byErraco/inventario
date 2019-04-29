from decimal import getcontext, Decimal

from openpyxl.styles import Alignment

from django.shortcuts import render
from django.http import Http404, HttpResponse
from django.db.models import Q
from django.views import View
from django.forms import formset_factory
from django.utils.timezone import localtime

from inventario.views import ListView, ReportView
from almacenes.models import Almacen
from productos.models import Producto 

from inventario.views import CreateUpdateListView

from .forms import CompraProductoForm, CostoCompraFormset, TrasladoProductoForm, AjusteInventarioProductoForm, FabricacionForm, FabricacionLoteForm, FabricacionLoteBaseFormSet, ConfirmarTrasladoProductoForm, EliminarTrasladoProductoForm, EliminarCompraProductoForm, EliminarFabricacionForm, EliminarAjusteInventarioProductoForm
from .filters import VentaProductoFilter, CompraProductoFilter, AjusteInventarioProductoFilter, FabricacionFilter, TrasladoProductoFilter

# Create your views here.


class VentasProductoView(ListView):
    permission_required = 'movimientos.view_ventaproducto'
    template_name = 'base_lista.html'
    template_table = 'movimientos/tabla_body_venta.html'
    paginate_by = 25
    extra_context = {'modelo': 'venta', 'titulo': 'Lista de ventas'}
    filter_ = VentaProductoFilter


class ComprasProductoView(CreateUpdateListView):
    view_permission = 'movimientos.view_compraproducto'
    add_permission = 'movimientos.add_compraproducto'
    change_permission = 'movimientos.add_compraproducto'
    template_name = 'base_lista_crear.html'
    extra_context = {'modelo': 'compra', 'titulo': 'Compras', 'titulo_form': 'Crear compra', 'boton_form': 'Crear'}
    template_table = 'movimientos/tabla_body_compra.html'
    paginate_by = 20
    form_class = CompraProductoForm
    filter_ = CompraProductoFilter

    def get_object_forms(self, compra_producto):
        forms = {}

        if self.request.user.has_almacen_perm(compra_producto.unidad_inventario.control_stock.almacen, 'delete_compraproducto'):
            forms['Eliminar'] = EliminarCompraProductoForm(initial={'object_id':compra_producto.id})
        return forms


class CompraProductoCreateUpdateView(View):
    template_name = 'forms.html'
    created_message = "La compra fue creada"

    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            if EliminarCompraProductoForm.nombre_form in request.POST:
                form = EliminarCompraProductoForm(request.POST or None)

                if form.is_valid():
                    form.delete()

                    return HttpResponse('Compra eliminada')

                return HttpResponse('No se pudo eliminar la compra')
            else:
                context = { 'forms': [], 'titulo_form': 'crear compra'}

                form = CompraProductoForm(request.POST or None)
                formset = CostoCompraFormset(request.POST)

                if form.is_valid() and formset.is_valid():
                    new_object = form.save()

                    formset = CostoCompraFormset(request.POST, instance=new_object)
                    formset.save()

                    context['forms'].append(CompraProductoForm())
                    context['forms'].append(CostoCompraFormset())

                    context['success'] = self.created_message

                    return render(request, self.template_name, context)

                context['forms'].append(form)
                context['forms'].append(formset)
                return render(request, self.template_name, context)

        raise Http404


def get_initial(querydict):
    initial = {}
    for key, item in querydict.lists():
        if key.endswith('[]'):
            initial[key] = item
        else:
            initial[key] = item[0]
    return initial


class CompraProductoFormView(View):
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            forms = []

            form = CompraProductoForm(initial = get_initial(request.POST))
            forms.append(form)

            prefix = CostoCompraFormset().prefix
            totalForms = request.POST.get(prefix+'-'+'TOTAL_FORMS')
            initial = []
            if totalForms:
                for i in range(int(totalForms)):
                    subformPrefix = prefix+'-'+str(i)+'-'
                    subformInitial = {key.replace(subformPrefix, ''):value for key, value in request.POST.items() if subformPrefix in key}
                    initial.append(subformInitial)

            form = CostoCompraFormset(initial=initial)
            forms.append(form)

            return render(request, 'forms.html', {'forms': forms, 'titulo_form': 'crear compra'})
            
        raise Http404


class TrasladosProductoView(CreateUpdateListView):
    view_permission = 'movimientos.view_trasladoproducto'
    add_permission = 'movimientos.add_trasladoproducto'
    change_permission = 'movimientos.add_trasladoproducto'
    template_name = 'base_lista_crear.html'
    extra_context = {'modelo': 'traslado', 'titulo': 'Traslados', 'titulo_form': 'Crear traslado', 'boton_form': 'Crear', 'url_icono': '/img/iconos/48/crear-traslado.png'}
    template_table = 'movimientos/tabla_body_traslado.html'
    paginate_by = 20
    form_class = TrasladoProductoForm
    filter_ = TrasladoProductoFilter

    def get_object_forms(self, traslado):
        user = self.request.user
        forms = {}
        if traslado.fecha_confirmacion is None:
            if user.has_almacen_perm(traslado.unidad_inventario_destino.control_stock.almacen, 'confirm_trasladoproducto'):
                forms['Confirmar'] = ConfirmarTrasladoProductoForm(initial={'object_id':traslado.id})
            if user.has_almacen_perm(traslado.unidad_inventario_origen.control_stock.almacen, 'delete_trasladoproducto'):
                forms['Eliminar'] = EliminarTrasladoProductoForm(initial={'object_id':traslado.id})
        else:
            if user.has_almacen_perm(traslado.unidad_inventario_destino.control_stock.almacen, 'delete_trasladoproducto'):
                forms['Eliminar'] = EliminarTrasladoProductoForm(initial={'object_id':traslado.id})
        return forms


class TrasladoProductoCreateUpdateView(View):
    template_name = 'forms.html'
    created_message = "El traslado fue creado"

    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            context = { 'forms': [], 'titulo_form': 'crear traslado'}
            
            if ConfirmarTrasladoProductoForm.nombre_form in request.POST:
                form = ConfirmarTrasladoProductoForm(request.POST or None)

                if form.is_valid():
                    form.save()

                    return HttpResponse('Traslado confirmado')

                return HttpResponse('No se pudo confirmar el traslado')
            elif EliminarTrasladoProductoForm.nombre_form in request.POST:
                form = EliminarTrasladoProductoForm(request.POST or None)

                if form.is_valid():
                    form.delete()

                    return HttpResponse('Traslado eliminado')

                return HttpResponse('No se pudo eliminar el traslado')
            else:
                form = TrasladoProductoForm(request.POST or None)

                if form.is_valid():
                    form.save()

                    context['forms'].append(TrasladoProductoForm())

                    context['success'] = self.created_message

                    return render(request, self.template_name, context)

                context['forms'].append(form)
                return render(request, self.template_name, context)

        raise Http404


class TrasladoProductoFormView(View):
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            forms = []
            form = TrasladoProductoForm(initial = get_initial(request.POST))
            forms.append(form)
            return render(request, 'forms.html', {'forms': forms, 'titulo_form': 'crear traslado'})
            
        raise Http404


class AjustesInventarioProductoView(CreateUpdateListView):
    view_permission = 'movimientos.view_ajusteinventarioproducto'
    add_permission = 'movimientos.add_ajusteinventarioproducto'
    change_permission = 'movimientos.add_ajusteinventarioproducto'
    template_name = 'base_lista_crear.html'
    extra_context = {'modelo': 'ajuste', 'titulo': 'Ajustes', 'titulo_form': 'Crear ajuste', 'boton_form': 'Crear'}
    template_table = 'movimientos/tabla_body_ajuste.html'
    paginate_by = 20
    form_class = AjusteInventarioProductoForm
    filter_ = AjusteInventarioProductoFilter

    def get_object_forms(self, ajuste):
        forms = {}

        if ajuste.fabricacion is None:
            if self.request.user.has_almacen_perm(ajuste.unidad_inventario.control_stock.almacen, 'delete_ajusteinventarioproducto'):
                forms['Eliminar'] = EliminarAjusteInventarioProductoForm(initial={'object_id':ajuste.id})
        return forms


class AjusteInventarioProductoCreateView(View):
    template_name = 'forms.html'
    created_message = "El ajuste fue creado"

    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            if EliminarAjusteInventarioProductoForm.nombre_form in request.POST:
                form = EliminarAjusteInventarioProductoForm(request.POST or None)

                if form.is_valid():
                    form.delete()

                    return HttpResponse('Ajuste eliminado')

                return HttpResponse('No se pudo eliminar el ajuste')
            else:
                context = { 'forms': [], 'titulo_form': 'crear ajuste'}

                form = AjusteInventarioProductoForm(request.POST or None)

                if form.is_valid():
                    form.save()

                    context['forms'].append(AjusteInventarioProductoForm())

                    context['success'] = self.created_message

                    return render(request, self.template_name, context)

                context['forms'].append(form)
                return render(request, self.template_name, context)

        raise Http404


class AjusteInventarioProductoFormView(View):
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            forms = []
            form = AjusteInventarioProductoForm(initial = get_initial(request.POST))
            forms.append(form)
            return render(request, 'forms.html', {'forms': forms, 'titulo_form': 'crear ajuste'})
            
        raise Http404


class FabricacionView(CreateUpdateListView):
    view_permission = 'movimientos.view_fabricacion'
    add_permission = 'movimientos.add_fabricacion'
    change_permission = 'movimientos.add_fabricacion'
    template_name = 'base_lista_crear.html'
    extra_context = {'modelo': 'fabricacion', 'titulo': 'Fabricación', 'titulo_form': 'Crear fabricación', 'boton_form': 'Crear'}
    template_table = 'movimientos/tabla_body_fabricacion.html'
    paginate_by = 20
    form_class = FabricacionForm
    filter_ = FabricacionFilter

    def get_object_forms(self, fabricacion):
        forms = {}

        if self.request.user.has_almacen_perm(fabricacion.unidad_inventario.control_stock.almacen, 'delete_fabricacion'):
            forms['Eliminar'] = EliminarFabricacionForm(initial={'object_id':fabricacion.id})
        return forms


class FabricacionCreateView(View):
    template_name = 'forms.html'
    created_message = "Los productos fueron creados"

    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            if EliminarFabricacionForm.nombre_form in request.POST:
                form = EliminarFabricacionForm(request.POST or None)

                if form.is_valid():
                    form.delete()

                    return HttpResponse('Fabricación eliminada')

                return HttpResponse('No se pudo eliminar la fabricación')
            else:
                context = { 'forms': [], 'titulo_form': 'crear fabricación' }

                form = FabricacionForm(request.POST or None)
                forms = []

                cantidad_produccion = request.POST.get('cantidad_produccion')
                if cantidad_produccion:
                    almacen_id = request.POST.get('almacen_produccion')
                    producto_id = request.POST.get('producto')
                    subforms_initials = {campo:valor for campo, valor in request.POST.items() if campo.startswith('componente_')}

                    producto = Producto.objects.actuales().get(id = producto_id)
                    almacen = Almacen.objects.actuales().get(id = almacen_id)

                    for componente_id, cantidad_componente in producto.componentes_totales(Q(producto_componente__seguimiento__isnull=False)).items():
                        componente = Producto.objects.get(id=componente_id)
                        prefix = 'componente_'+str(componente_id)
                        totalForms = subforms_initials.get(prefix+'-'+'TOTAL_FORMS')
                        initial = []
                        if totalForms:
                            for i in range(int(totalForms)):
                                subformPrefix = prefix+'-'+str(i)+'-'
                                subformInitial = {key.replace(subformPrefix, ''):value for key, value in subforms_initials.items() if subformPrefix in key}
                                initial.append(subformInitial)

                        opciones_totales = almacen.codigos_lotes_producto(componente)
                        seleccionados = []
                        
                        for subform_initial in initial:
                            seleccionados.append(subform_initial['codigo_lote'])

                        opciones = [opcion for opcion in opciones_totales if opcion[0] not in seleccionados]

                        getcontext().prec = 4

                        cantidad_componente = cantidad_componente*Decimal(cantidad_produccion)
                        FabricacionLoteFormset = formset_factory(form=FabricacionLoteForm, extra=0, max_num=len(opciones_totales), can_delete=False, formset=FabricacionLoteBaseFormSet)
                        formset = FabricacionLoteFormset(request.POST, prefix='componente_'+str(componente_id), form_kwargs={'almacen': almacen, 'componente': componente, 'cantidad_componente': cantidad_componente, 'opciones':opciones}, initial=initial)
                        forms.append(formset)


                if form.is_valid():
                    valido = True
                    for formset in forms:
                        if not formset.is_valid():
                            valido = False

                    if valido:
                        fabricacion = form.save()

                        for formset in forms:
                            formset.save(fabricacion)

                        context['forms'].append(FabricacionForm())

                        context['success'] = self.created_message

                        return render(request, self.template_name, context)

                context['forms'].append(form)
                for form in forms:
                    context['forms'].append(form)
                return render(request, self.template_name, context)

        raise Http404


class FabricacionFormView(View):
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            forms = []
            form = FabricacionForm(initial = dict(request.POST.items()))
            forms.append(form)

            getcontext().prec = 4
            
            cantidad_produccion = request.POST.get('cantidad_produccion')
            if cantidad_produccion:
                almacen_id = request.POST.get('almacen_produccion')
                changed_data = request.POST.get('changed_data')
                agregar_componente = ''

                if not changed_data:    
                    producto_id = request.POST.get('producto')
                    subforms_initials = {campo:valor for campo, valor in request.POST.items() if campo.startswith('componente_')}
                elif changed_data != 'almacen_produccion':
                    producto_id = request.POST.get('producto')

                    if changed_data != 'producto':
                        subforms_initials = {campo:valor for campo, valor in request.POST.items() if campo.startswith('componente_')}
                        
                        if changed_data.endswith('añadir'):
                            agregar_componente = changed_data.split('-')[0]
                    else:
                        subforms_initials = []
                else:
                    producto_id = None
                    subforms_initials = []

                if almacen_id and producto_id:
                    producto = Producto.objects.actuales().get(id = producto_id)
                    almacen = Almacen.objects.actuales().get(id = almacen_id)

                    if subforms_initials:
                        for componente_id, cantidad_componente in producto.componentes_totales(Q(producto_componente__seguimiento__isnull=False)).items():
                            componente = Producto.objects.get(id=componente_id)
                            prefix = 'componente_'+str(componente_id)
                            totalForms = subforms_initials.get(prefix+'-'+'TOTAL_FORMS')
                            initial = []
                            if totalForms:
                                for i in range(int(totalForms)):
                                    subformPrefix = prefix+'-'+str(i)+'-'
                                    subformInitial = {key.replace(subformPrefix, ''):value for key, value in subforms_initials.items() if subformPrefix in key}
                                    initial.append(subformInitial)
                            opciones_totales = almacen.codigos_lotes_producto(componente)
                            seleccionados = []
                            
                            for subform_initial in initial:
                                seleccionados.append(subform_initial['codigo_lote'])

                            opciones = [opcion for opcion in opciones_totales if opcion[0] not in seleccionados]

                            if agregar_componente == prefix:
                                initial.append({'codigo_lote': opciones.pop(0)[0]})

                            cantidad_componente = cantidad_componente*Decimal(cantidad_produccion)
                            FabricacionLoteFormset = formset_factory(form=FabricacionLoteForm, extra=0, max_num=len(opciones_totales), can_delete=False, formset=FabricacionLoteBaseFormSet)
                            formset = FabricacionLoteFormset(prefix='componente_'+str(componente_id), form_kwargs={'almacen': almacen, 'componente': componente, 'cantidad_componente': cantidad_componente, 'opciones':opciones}, initial=initial)
                            forms.append(formset)
                    else:
                        for componente_id, cantidad_componente in producto.componentes_totales(Q(producto_componente__seguimiento__isnull=False)).items():
                            componente = Producto.objects.get(id=componente_id)
                            opciones = almacen.codigos_lotes_producto(componente)
                            cantidad_componente = cantidad_componente*Decimal(cantidad_produccion)
                            FabricacionLoteFormset = formset_factory(form=FabricacionLoteForm, extra=1, max_num=len(opciones), can_delete=False, formset=FabricacionLoteBaseFormSet)
                            formset = FabricacionLoteFormset(prefix='componente_'+str(componente_id), form_kwargs={'almacen': almacen, 'componente': componente, 'cantidad_componente': cantidad_componente, 'opciones':opciones})
                            forms.append(formset)

            return render(request, 'forms.html', {'forms': forms, 'titulo_form': 'crear fabricación'})
            
        raise Http404


class ComprasReportView(ReportView):
    filter_ = CompraProductoFilter
    filename = 'reporte_compras'
    title = 'Reporte de compras'

    def append_objects(self, request, ws, **kwargs):
        filter_ = self.filter_(request.GET, request=request, **kwargs)
        
        self.append_filter(filter_, ws)
        
        ws.append(['Almacén', 'Producto', 'Código de lote/serie', 'Proveedor', 'Cantidad', 'Costo por unidad', 'Costos adicionales','','Fecha de creación'])
        ws.merge_cells(start_row=ws.max_row, start_column=7, end_row=ws.max_row, end_column=8)

        object_list = filter_.qs

        for obj in object_list:
            max_row = ws.max_row+1
            ws.cell(row=max_row, column=1).value = str(obj.unidad_inventario.control_stock.almacen)
            ws.cell(row=max_row, column=2).value = str(obj.unidad_inventario.control_stock.producto)
            if obj.unidad_inventario.lote_produccion:
                ws.cell(row=max_row, column=3).value = str(obj.unidad_inventario.lote_produccion.codigo)

            ws.cell(row=max_row, column=4).value = str(obj.proveedor)
            ws.cell(row=max_row, column=5).value = obj.cantidad
            ws.cell(row=max_row, column=6).value = obj.costo_unidad

            row = max_row
            for costo in obj.costos.all():
                ws.cell(row=row, column=7).value = costo.descripcion+':'
                ws.cell(row=row, column=8).value = costo.cantidad
                row = ws.max_row+1

            if row > max_row+1:
                for column in [1,2,3,4,5,6,9,10]:
                    ws.merge_cells(start_row=max_row, start_column=column, end_row=row-1, end_column=column)
                    ws.cell(row=max_row, column=column).alignment = Alignment(vertical='top')

            ws.cell(row=max_row, column=9).value = localtime(obj.fecha_creacion).strftime('%m-%d-%Y %I:%M %p')


class AjustesReportView(ReportView):
    filter_ = AjusteInventarioProductoFilter
    filename = 'reporte_ajustes'
    title = 'Reporte de ajustes'

    def append_objects(self, request, ws, **kwargs):
        filter_ = self.filter_(request.GET, request=request, **kwargs)
        
        self.append_filter(filter_, ws)

        ws.append(['Almacén', 'Producto', 'Código de lote/serie', 'Cantidad', 'Descripción', 'Fecha de creación'])

        object_list = filter_.qs

        for obj in object_list:
            max_row = ws.max_row+1
            ws.cell(row=max_row, column=1).value = str(obj.unidad_inventario.control_stock.almacen)
            ws.cell(row=max_row, column=2).value = str(obj.unidad_inventario.control_stock.producto)
            if obj.unidad_inventario.lote_produccion:
                ws.cell(row=max_row, column=3).value = str(obj.unidad_inventario.lote_produccion.codigo)

            ws.cell(row=max_row, column=4).value = obj.cantidad

            if obj.descripcion:
                ws.cell(row=max_row, column=5).value = obj.descripcion
            else:
                ws.cell(row=max_row, column=5).value = 'Ajuste'

            ws.cell(row=max_row, column=6).value = localtime(obj.fecha_creacion).strftime('%m-%d-%Y %I:%M %p')


class VentasReportView(ReportView):
    filter_ = VentaProductoFilter
    filename = 'reporte_ventas'
    title = 'Reporte de ventas'

    def append_objects(self, request, ws, **kwargs):
        filter_ = self.filter_(request.GET, request=request, **kwargs)
        
        self.append_filter(filter_, ws)

        ws.append(['Almacén', 'Producto', 'Código de lote/serie', 'Cantidad', 'Id de venta', 'Estado de venta', 'Fecha de creación'])

        object_list = filter_.qs

        for obj in object_list:
            max_row = ws.max_row+1
            ws.cell(row=max_row, column=1).value = str(obj.unidad_inventario.control_stock.almacen)
            ws.cell(row=max_row, column=2).value = str(obj.unidad_inventario.control_stock.producto)
            if obj.unidad_inventario.lote_produccion:
                ws.cell(row=max_row, column=3).value = str(obj.unidad_inventario.lote_produccion.codigo)

            ws.cell(row=max_row, column=4).value = obj.cantidad
            ws.cell(row=max_row, column=5).value = obj.venta.id
            ws.cell(row=max_row, column=6).value = obj.venta.estado_str().capitalize()
            ws.cell(row=max_row, column=7).value = localtime(obj.fecha_creacion).strftime('%m-%d-%Y %I:%M %p')


class FabricacionReportView(ReportView):
    filter_ = FabricacionFilter
    filename = 'reporte_fabricacion'
    title = 'Reporte de fabricación'

    def append_objects(self, request, ws, **kwargs):
        filter_ = self.filter_(request.GET, request=request, **kwargs)
        
        self.append_filter(filter_, ws)

        ws.append(['Código de lote/serie', 'Almacén', 'Producto', 'Cantidad', 'Fecha de producción', 'Fecha de vencimiento', 'Fecha de creación'])

        object_list = filter_.qs

        for obj in object_list:
            max_row = ws.max_row+1

            ws.cell(row=max_row, column=1).value = obj.lote_produccion.codigo
            ws.cell(row=max_row, column=2).value = str(obj.unidad_inventario.control_stock.almacen)
            ws.cell(row=max_row, column=3).value = str(obj.lote_produccion.producto)
            ws.cell(row=max_row, column=4).value = obj.cantidad_produccion
            ws.cell(row=max_row, column=5).value = localtime(obj.lote_produccion.fecha_produccion).strftime('%m-%d-%Y %I:%M %p')

            if obj.lote_produccion.fecha_vencimiento:
                ws.cell(row=max_row, column=6).value = localtime(obj.lote_produccion.fecha_vencimiento).strftime('%m-%d-%Y %I:%M %p')

            ws.cell(row=max_row, column=7).value = localtime(obj.fecha_creacion).strftime('%m-%d-%Y %I:%M %p')


class TrasladosReportView(ReportView):
    filter_ = TrasladoProductoFilter
    filename = 'reporte_traslados'
    title = 'Reporte de traslados'

    def append_objects(self, request, ws, **kwargs):
        filter_ = self.filter_(request.GET, request=request, **kwargs)

        self.append_filter(filter_, ws)

        ws.append(['Almacén origen', 'Producto', 'Código de lote/serie', 'Cantidad', 'Costo', 'Almacén destino', 'Fecha de confirmación', 'Fecha de creación'])

        object_list = filter_.qs

        for obj in object_list:
            max_row = ws.max_row+1

            ws.cell(row=max_row, column=1).value = str(obj.unidad_inventario_origen.control_stock.almacen)
            ws.cell(row=max_row, column=2).value = str(obj.unidad_inventario_origen.control_stock.producto)

            if obj.unidad_inventario_origen.lote_produccion:
                ws.cell(row=max_row, column=3).value = str(obj.unidad_inventario_origen.lote_produccion.codigo)

            ws.cell(row=max_row, column=4).value = obj.cantidad

            if obj.costo:
                ws.cell(row=max_row, column=5).value = obj.costo

            ws.cell(row=max_row, column=6).value = str(obj.unidad_inventario_destino.control_stock.almacen)

            if obj.fecha_confirmacion:
                ws.cell(row=max_row, column=7).value = localtime(obj.fecha_confirmacion).strftime('%m-%d-%Y %I:%M %p')

            ws.cell(row=max_row, column=8).value = localtime(obj.fecha_creacion).strftime('%m-%d-%Y %I:%M %p')